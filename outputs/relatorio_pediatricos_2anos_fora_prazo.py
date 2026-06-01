#!/usr/bin/env python3
"""Relatório: pacientes com até 2 anos de IDADE cujo retorno semestral está vencido.

REGRA CLÍNICA: oftalmopediatria de bebês/crianças ≤ 2 anos exige acompanhamento
a cada 6 meses. Quem passou desse prazo sem nova consulta precisa ser
recontatado pra preservar o acompanhamento.

Como rodar no Mac:
    cd ~/Documents/Claude/Projects/AGENTE\\ IA\\ BLINK
    pip3 install --user openpyxl httpx
    python3 outputs/relatorio_pediatricos_2anos_fora_prazo.py

Lê credenciais Medware e Kommo do .env do voice_agent (mesmas usadas em prod).

Saída:
    outputs/RELATORIO_pediatricos_fora_prazo_DD-MM-YYYY.xlsx

Filtros aplicados (todos têm que bater):
    - paciente nascido em ou após 01/06/2024 (idade ≤ 2 anos)
    - última consulta REALIZADA (status=5) > 6 meses atrás
    - sem agendamento FUTURO com status 1 (agendado), 2 (confirmar), 3 (confirmado)
"""
from __future__ import annotations

import json
import os
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

# Garantir que voice_agent é importável (rodando do root do repo)
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

# ---------- Carrega .env do projeto ----------
def carregar_env():
    env_path = ROOT / ".env"
    if not env_path.exists():
        env_path = ROOT / "voice_agent" / ".env"
    if env_path.exists():
        for linha in env_path.read_text(encoding="utf-8").splitlines():
            linha = linha.strip()
            if not linha or linha.startswith("#"):
                continue
            if "=" in linha:
                k, _, v = linha.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

carregar_env()

try:
    import httpx
except ImportError:
    print("ERRO: instale httpx: pip3 install --user httpx", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("ERRO: instale openpyxl: pip3 install --user openpyxl", file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# Constantes — Medware
# ---------------------------------------------------------------------------

MEDWARE_BASE = os.environ.get(
    "MEDWARE_BASE_URL", "https://medware.blinkoftalmologia.com.br/api",
).rstrip("/")
MEDWARE_USER = os.environ.get("MEDWARE_USERNAME") or os.environ.get("MEDWARE_USER")
MEDWARE_PASS = os.environ.get("MEDWARE_PASSWORD") or os.environ.get("MEDWARE_SENHA")

STATUS_REALIZADO = 5
STATUS_FUTUROS = {1, 2, 3}  # agendado, confirmar, confirmado

HOJE = datetime.now()
LIMITE_IDADE = HOJE - timedelta(days=2 * 365)         # nascidos depois disso
LIMITE_PRAZO_6M = HOJE - timedelta(days=180)          # consultas antes disso = vencidas
PERIODO_BUSCA_INI = (HOJE - timedelta(days=730)).strftime("%d/%m/%Y")  # 24 meses
PERIODO_BUSCA_FIM = (HOJE + timedelta(days=180)).strftime("%d/%m/%Y")  # inclui futuros 6m

# Procedimento conhecido = pediatria menos de 3 anos
PROCEDIMENTOS_PEDIATRICOS = {311}


# ---------------------------------------------------------------------------
# Cliente Medware (mesmo padrão do voice_agent/medware.py)
# ---------------------------------------------------------------------------

class MedwareClient:
    def __init__(self):
        if not MEDWARE_USER or not MEDWARE_PASS:
            print("ERRO: MEDWARE_USERNAME/MEDWARE_PASSWORD ausentes no .env",
                  file=sys.stderr)
            sys.exit(3)
        self._token = None
        self._token_exp = 0
        self._cli = httpx.Client(timeout=30.0)

    def _login(self):
        r = self._cli.post(
            f"{MEDWARE_BASE}/Acesso/login",
            json={"identificacao": MEDWARE_USER, "senha": MEDWARE_PASS},
        )
        r.raise_for_status()
        data = r.json()
        self._token = data.get("token") or data.get("Token")
        if not self._token:
            print(f"ERRO: token não veio no login: {data}", file=sys.stderr)
            sys.exit(4)
        # Token JWT vale ~24h
        self._token_exp = time.time() + 23 * 3600
        print("✓ Login Medware OK", file=sys.stderr)

    def _headers(self):
        if not self._token or time.time() > self._token_exp:
            self._login()
        return {"Authorization": f"Bearer {self._token}"}

    def listar_agendamentos(self, data_inicio: str, data_fim: str) -> list[dict]:
        params = {"dataInicio": data_inicio, "dataFim": data_fim}
        r = self._cli.get(
            f"{MEDWARE_BASE}/Agendamento/Listar",
            params=params, headers=self._headers(),
        )
        if r.status_code != 200:
            print(f"ERRO listar_agendamentos {data_inicio}-{data_fim}: "
                  f"{r.status_code} {r.text[:300]}", file=sys.stderr)
            return []
        return r.json() or []

    def buscar_paciente(self, cod_paciente: int) -> dict | None:
        r = self._cli.get(
            f"{MEDWARE_BASE}/Pacientes/buscar",
            params={"codPaciente": cod_paciente},
            headers=self._headers(),
        )
        if r.status_code != 200:
            return None
        data = r.json() or []
        if isinstance(data, list):
            return data[0] if data else None
        return data


# ---------------------------------------------------------------------------
# Cliente Kommo — só pra buscar URL do lead por telefone
# ---------------------------------------------------------------------------

KOMMO_SUBDOMAIN = os.environ.get("KOMMO_SUBDOMAIN", "univeja")
KOMMO_TOKEN = os.environ.get("KOMMO_LONG_LIVED_TOKEN") or os.environ.get("KOMMO_TOKEN")


def buscar_url_kommo(telefone: str, cli: httpx.Client) -> str | None:
    """Procura lead Kommo por telefone. Retorna URL ou None."""
    if not telefone or not KOMMO_TOKEN:
        return None
    digitos = "".join(c for c in telefone if c.isdigit())
    if len(digitos) < 8:
        return None
    # Pega últimos 8 dígitos (núcleo do número)
    chave = digitos[-9:] if len(digitos) >= 9 else digitos
    try:
        r = cli.get(
            f"https://{KOMMO_SUBDOMAIN}.kommo.com/api/v4/leads",
            params={"query": chave, "limit": 1},
            headers={"Authorization": f"Bearer {KOMMO_TOKEN}"},
        )
        if r.status_code != 200:
            return None
        data = r.json() or {}
        leads = (data.get("_embedded") or {}).get("leads") or []
        if not leads:
            return None
        lid = leads[0].get("id")
        return f"https://{KOMMO_SUBDOMAIN}.kommo.com/leads/detail/{lid}"
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------------

def calcular_idade_anos(data_nasc_iso: str) -> float | None:
    """Retorna idade em anos com decimal. None se data inválida."""
    if not data_nasc_iso:
        return None
    try:
        # Aceita "2024-06-01" ou "2024-06-01T00:00:00"
        d = datetime.strptime(data_nasc_iso[:10], "%Y-%m-%d")
        delta = HOJE - d
        return round(delta.days / 365.25, 2)
    except Exception:
        return None


def main():
    print(f"Hoje: {HOJE.strftime('%d/%m/%Y')}")
    print(f"Período Medware varredura: {PERIODO_BUSCA_INI} → {PERIODO_BUSCA_FIM}")
    print(f"Limite idade: nascidos após {LIMITE_IDADE.strftime('%d/%m/%Y')}")
    print(f"Prazo vencido: última consulta antes de {LIMITE_PRAZO_6M.strftime('%d/%m/%Y')}")
    print()

    mw = MedwareClient()

    # 1. Varre Medware em chunks de 3 meses pra não estourar
    todos_agendamentos = []
    cursor = HOJE - timedelta(days=730)
    fim_global = HOJE + timedelta(days=180)
    while cursor < fim_global:
        prox = cursor + timedelta(days=90)
        if prox > fim_global:
            prox = fim_global
        ini = cursor.strftime("%d/%m/%Y")
        fim = prox.strftime("%d/%m/%Y")
        print(f"  buscando {ini} → {fim} ...")
        lote = mw.listar_agendamentos(ini, fim)
        print(f"     +{len(lote)} agendamentos")
        todos_agendamentos.extend(lote)
        cursor = prox + timedelta(days=1)

    print(f"\nTotal de agendamentos brutos: {len(todos_agendamentos)}")

    # 2. Agrupa por paciente
    consultas_por_paciente: dict[int, list[dict]] = defaultdict(list)
    for ag in todos_agendamentos:
        cod_p = ag.get("codPaciente") or ag.get("codpaciente")
        if cod_p:
            consultas_por_paciente[int(cod_p)].append(ag)

    print(f"Pacientes únicos: {len(consultas_por_paciente)}")

    # 3. Pra cada paciente: descobrir data nascimento + última consulta realizada + sem agend futuro
    candidatos = []
    cli_kommo = httpx.Client(timeout=15.0)
    for i, (cod_p, lista) in enumerate(consultas_por_paciente.items()):
        if (i + 1) % 50 == 0:
            print(f"  processando {i + 1}/{len(consultas_por_paciente)}...")

        # Filtra realizadas e agendadas
        realizadas = [
            a for a in lista
            if (a.get("codStatusAgendamento") or a.get("status")) == STATUS_REALIZADO
        ]
        futuras_pendentes = [
            a for a in lista
            if (a.get("codStatusAgendamento") or a.get("status")) in STATUS_FUTUROS
            and datetime.strptime((a.get("data") or a.get("dataAgendamento") or "1900-01-01")[:10],
                                  "%Y-%m-%d") >= HOJE
        ]

        if not realizadas:
            continue
        if futuras_pendentes:
            # Já tem consulta futura — não está fora do prazo
            continue

        # Última realizada
        def _data_ag(a):
            return (a.get("data") or a.get("dataAgendamento") or "1900-01-01")[:10]

        realizadas.sort(key=_data_ag, reverse=True)
        ultima = realizadas[0]
        data_ultima = datetime.strptime(_data_ag(ultima), "%Y-%m-%d")

        if data_ultima > LIMITE_PRAZO_6M:
            # Última consulta há menos de 6 meses → ainda dentro do prazo
            continue

        # Busca dados do paciente (idade + telefone)
        paciente = mw.buscar_paciente(cod_p)
        if not paciente:
            continue
        data_nasc = paciente.get("dataNasc") or paciente.get("dataNascimento") or ""
        idade = calcular_idade_anos(data_nasc)
        if idade is None or idade > 2.0:
            continue

        telefone = (
            paciente.get("telefonePaciente")
            or paciente.get("telefone")
            or paciente.get("celular")
            or ""
        )
        nome = paciente.get("nomePaciente") or paciente.get("nome") or "?"

        kommo_url = buscar_url_kommo(telefone, cli_kommo) if telefone else None

        candidatos.append({
            "codPaciente": cod_p,
            "nome": nome,
            "dataNasc": data_nasc[:10] if data_nasc else "",
            "idadeAnos": idade,
            "telefone": telefone,
            "ultimaConsulta": data_ultima.strftime("%d/%m/%Y"),
            "diasDesdeUltima": (HOJE - data_ultima).days,
            "mesesDesdeUltima": round((HOJE - data_ultima).days / 30, 1),
            "medicoUltima": ultima.get("nomeMedico") or "?",
            "kommoUrl": kommo_url or "(não encontrado)",
        })

    candidatos.sort(key=lambda c: c["diasDesdeUltima"], reverse=True)
    print(f"\n✓ {len(candidatos)} pacientes pediátricos ≤ 2 anos fora do prazo de 6m")

    # 4. Gera Excel
    out_path = ROOT / "outputs" / (
        f"RELATORIO_pediatricos_fora_prazo_{HOJE.strftime('%d-%m-%Y')}.xlsx"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Pediátricos fora do prazo"

    cabecalho = [
        "Cód Paciente", "Nome", "Data Nasc", "Idade (anos)", "Telefone",
        "Última Consulta", "Dias desde", "Meses desde",
        "Médico", "URL Kommo",
    ]
    ws.append(cabecalho)

    azul = PatternFill("solid", fgColor="1F4E78")
    branco = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = azul
        cell.font = branco
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in candidatos:
        ws.append([
            c["codPaciente"], c["nome"], c["dataNasc"], c["idadeAnos"],
            c["telefone"], c["ultimaConsulta"], c["diasDesdeUltima"],
            c["mesesDesdeUltima"], c["medicoUltima"], c["kommoUrl"],
        ])

    larguras = [12, 30, 12, 10, 16, 14, 10, 11, 22, 50]
    for col, larg in zip("ABCDEFGHIJ", larguras):
        ws.column_dimensions[col].width = larg

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"\n✓ Excel gerado: {out_path}")
    print(f"  {len(candidatos)} linhas")
    print("\nAbra com: open '{out_path}'")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrompido (Ctrl+C).")
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ ERRO: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
